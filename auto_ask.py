import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import Playwright, sync_playwright
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError

questions = [
    "I heard the Z Flip8 just came out, what are its key features?",
    "질문 2",
    "질문 3",
    "질문 4",
]

OUTPUT_DIR = Path(r"\\내부IP주소\ict혁신그룹2팀_se통합검증")  # 실제 내부 IP로 교체

results = []


def click_if_visible(locator, timeout: int = 3000) -> bool:
    """locator가 timeout 안에 나타나면 클릭하고 True, 안 나타나면 스킵하고 False"""
    try:
        locator.wait_for(state="visible", timeout=timeout)
        locator.click()
        return True
    except PlaywrightTimeoutError:
        return False


def open_chat_if_needed(page) -> None:
    """Open chat 버튼이 뜨면 클릭, 이미 챗봇 화면이 노출돼 있으면 그냥 넘어감"""
    click_if_visible(page.get_by_role("button", name="Open chat"), timeout=5000)


def ask_and_get_answer(page, question: str) -> str:
    page.get_by_role("textbox", name="Ask me anything").click()
    page.get_by_role("textbox", name="Ask me anything").fill(question)
    page.get_by_role("button", name="Send message").click()

    # 답변이 완료되면 피드백 툴팁(👍/👎 등)이 노출됨 -> 그걸 완료 신호로 사용
    page.locator(".feedback-tolltip-wrapper").last.wait_for(state="visible", timeout=30000)

    return page.locator("div.chatbot-main-cont").last.inner_text()


def reset_chat(page) -> None:
    """채팅 닫고 다시 열어서 새 대화로 초기화 (상황에 따라 안 뜨는 버튼은 스킵)"""
    click_if_visible(page.get_by_role("button", name="Options"))
    click_if_visible(page.get_by_role("button", name="Close Chat Close Chat"))
    click_if_visible(page.get_by_role("button", name="Close").nth(2))
    click_if_visible(page.get_by_role("button").filter(has_text=re.compile(r"^$")).nth(5))
    open_chat_if_needed(page)
    click_if_visible(page.get_by_role("button", name="Cancel"))
    click_if_visible(page.get_by_role("button", name="Options"))
    click_if_visible(page.get_by_role("button", name="New Chat New Chat"))


def ensure_fresh_session(page, max_attempts: int = 3) -> None:
    """이전 대화 흔적이 남아있으면 reset_chat을 반복 시도해서 확실히 새 세션으로 만듦"""
    for attempt in range(max_attempts):
        if page.locator(".feedback-tolltip-wrapper").count() == 0:
            return
        reset_chat(page)
    raise RuntimeError("이전 세션 초기화에 반복적으로 실패했습니다.")


def run(playwright: Playwright) -> None:
    context = playwright.chromium.launch_persistent_context(
        user_data_dir="C:\\automation-profile",
        channel="chrome",
        headless=False,
    )
    page = context.new_page()
    page.goto("https://www.samsung.com/uk/")
    open_chat_if_needed(page)
    ensure_fresh_session(page)  # 이전 실행에서 남은 대화 세션 확실히 초기화

    for i, q in enumerate(questions):
        answer = ask_and_get_answer(page, q)
        results.append({
            "question": q,
            "answer": answer,
            "timestamp": datetime.now().isoformat(),
        })
        print(f"[{i + 1}/{len(questions)}] 완료: {q}")

        if i < len(questions) - 1:
            ensure_fresh_session(page)  # 리셋 시도 후 정말 깨끗해졌는지 확인, 안 됐으면 재시도

    # 최종 검증: question/answer가 빈 값인 항목이 있으면 재질의
    max_retry = 3
    for attempt in range(max_retry):
        blanks = [r for r in results if not r["question"].strip() or not r["answer"].strip()]
        if not blanks:
            break
        print(f"[재검증 {attempt + 1}/{max_retry}] 빈 값 {len(blanks)}건 재질의")
        for r in blanks:
            ensure_fresh_session(page)
            r["answer"] = ask_and_get_answer(page, r["question"])
            r["timestamp"] = datetime.now().isoformat()
    else:
        blanks = [r for r in results if not r["question"].strip() or not r["answer"].strip()]
        if blanks:
            raise RuntimeError(
                f"{max_retry}번 재시도했지만 여전히 빈 값이 있습니다: "
                f"{[r['question'] for r in blanks]}"
            )

    context.close()


with sync_playwright() as playwright:
    run(playwright)

df = pd.DataFrame(results)[["question", "answer"]].rename(
    columns={"question": "질문", "answer": "답변"}
)
excel_filename = OUTPUT_DIR / f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

FONT = Font(name="맑은 고딕", size=9)
WRAP = Alignment(wrap_text=True, vertical="top")
HEADER_FILL = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")

with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Sheet1")
    worksheet = writer.sheets["Sheet1"]

    worksheet.column_dimensions["A"].width = 70.7  # 약 500px
    worksheet.column_dimensions["B"].width = 85.0  # 약 600px

    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = FONT
            cell.alignment = WRAP

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
